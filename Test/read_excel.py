from openpyxl import load_workbook

# Read Vendor Details
print('=== VENDOR DETAILS ===')
wb = load_workbook('Vendor Details.xlsx')
ws = wb.active
print('Sheet name:', ws.title)
print('Max row:', ws.max_row)
print('Max column:', ws.max_column)
print('\nHeaders:')
headers = [cell.value for cell in ws[1]]
print(headers)
print('\nFirst 5 data rows:')
for i, row in enumerate(ws.iter_rows(min_row=2, max_row=6, values_only=True)):
    print(f'Row {i+2}: {row}')

print('\n' + '='*50)

# Read Vehicle and Loan Details
print('\n=== VEHICLE AND LOAN DETAILS ===')
wb2 = load_workbook('Vehicle and Loan Details.xlsx')
print('Sheet names:', wb2.sheetnames)
for sheet_name in wb2.sheetnames:
    ws2 = wb2[sheet_name]
    print(f'\n--- Sheet: {sheet_name} ---')
    print('Max row:', ws2.max_row)
    print('Max column:', ws2.max_column)
    if ws2.max_row > 0:
        headers = [cell.value for cell in ws2[1]]
        print('Headers:', headers)
        print('First 3 rows:')
        for i, row in enumerate(ws2.iter_rows(min_row=2, max_row=4, values_only=True)):
            print(f'  {row}')

print('\n' + '='*50)

# Read Rate List
print('\n=== ALL PARTY RATE FILE ===')
wb3 = load_workbook('ALL PARTY RATE FILE.xlsx')
print('Sheet names:', wb3.sheetnames)
for sheet_name in wb3.sheetnames[:3]:  # First 3 sheets
    ws3 = wb3[sheet_name]
    print(f'\n--- Sheet: {sheet_name} ---')
    print('Max row:', ws3.max_row)
    print('Max column:', ws3.max_column)
    if ws3.max_row > 0:
        headers = [cell.value for cell in ws3[1]]
        print('Headers:', headers)
        print('First 3 rows:')
        for i, row in enumerate(ws3.iter_rows(min_row=2, max_row=4, values_only=True)):
            print(f'  {row}')
