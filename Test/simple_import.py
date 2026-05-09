"""
Simple Import Script with direct file logging
"""
import sys
import os
from datetime import datetime
from openpyxl import load_workbook

# Redirect output to file
log_file = open('import_log.txt', 'w', encoding='utf-8')
original_stdout = sys.stdout
original_stderr = sys.stderr
sys.stdout = log_file
sys.stderr = log_file

print("Starting import...")
print(f"Time: {datetime.now()}")

# Add the project directory to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

try:
    from models import db, Vendor, Vehicle, Loan, Location, Tenant, app
    from webapp import generate_vendor_code
    print("Models imported successfully")
except Exception as e:
    print(f"Error importing models: {e}")
    import traceback
    traceback.print_exc()
    log_file.close()
    sys.exit(1)

def safe_str(value, max_len=None):
    if value is None:
        return None
    result = str(value).strip()
    if max_len and len(result) > max_len:
        result = result[:max_len]
    return result if result else None

def safe_float(value):
    if value is None:
        return 0.0
    try:
        if isinstance(value, str):
            value = value.replace(',', '').replace('₹', '').strip()
        return float(value)
    except:
        return 0.0

def parse_int(value):
    if value is None:
        return None
    try:
        return int(float(value))
    except:
        return None

def parse_date(date_value):
    if not date_value:
        return None
    if isinstance(date_value, datetime):
        return date_value.date() if hasattr(date_value, 'date') else date_value
    if isinstance(date_value, str):
        for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y']:
            try:
                return datetime.strptime(date_value, fmt).date()
            except:
                continue
    return None

def get_or_create_tenant():
    tenant = Tenant.query.first()
    if not tenant:
        tenant = Tenant(name="Default Tenant", subdomain="default")
        db.session.add(tenant)
        db.session.commit()
        print(f"Created tenant: {tenant.name} (ID: {tenant.id})")
    return tenant

def import_vendors(filepath, tenant_id):
    print(f"\n{'='*60}")
    print(f"Importing Vendors from: {filepath}")
    print(f"{'='*60}")
    
    wb = load_workbook(filepath)
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    print(f"Headers: {headers}")
    print(f"Total rows: {ws.max_row - 1}")
    
    imported = 0
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
        try:
            data = dict(zip(headers, row))
            
            vendor_name = safe_str(data.get('VENDOR NAME') or data.get('Company Name') or data.get('Vendor Name'))
            if not vendor_name:
                continue
            
            existing = Vendor.query.filter_by(tenant_id=tenant_id, vendor_name=vendor_name).first()
            if existing:
                print(f"  Skipping (exists): {vendor_name}")
                continue
            
            vendor_code = generate_vendor_code(tenant_id)
            
            vendor = Vendor(
                tenant_id=tenant_id,
                vendor_code=vendor_code,
                vendor_name=vendor_name,
                vendor_type=safe_str(data.get('TYPE') or data.get('Vendor Type'), 50) or 'supplier',
                status='active',
                contact_person=safe_str(data.get('CONTACT PERSON') or data.get('Contact Person'), 100),
                mobile=safe_str(data.get('MOBILE') or data.get('Mobile') or data.get('PHONE'), 20),
                email=safe_str(data.get('EMAIL') or data.get('Email'), 150),
                reg_city=safe_str(data.get('CITY') or data.get('City'), 100),
                reg_state=safe_str(data.get('STATE') or data.get('State'), 100),
                gstin=safe_str(data.get('GSTIN') or data.get('GST'), 20),
                pan=safe_str(data.get('PAN'), 20),
            )
            
            db.session.add(vendor)
            imported += 1
            print(f"  Imported: {vendor_name} (Code: {vendor_code})")
            
        except Exception as e:
            print(f"  ERROR Row {row_idx}: {e}")
    
    db.session.commit()
    print(f"\nVendors imported: {imported}")
    return imported

def import_vehicles(filepath, tenant_id):
    print(f"\n{'='*60}")
    print(f"Importing Vehicles from: {filepath}")
    print(f"{'='*60}")
    
    wb = load_workbook(filepath)
    vehicle_sheet = None
    
    for sheet_name in wb.sheetnames:
        if 'vehicle' in sheet_name.lower():
            vehicle_sheet = wb[sheet_name]
            break
    
    if not vehicle_sheet:
        vehicle_sheet = wb[wb.sheetnames[0]]
    
    headers = [cell.value for cell in vehicle_sheet[1]]
    print(f"Sheet: {vehicle_sheet.title}")
    print(f"Headers: {headers}")
    
    imported = 0
    
    for row_idx, row in enumerate(vehicle_sheet.iter_rows(min_row=2, max_row=vehicle_sheet.max_row, values_only=True), start=2):
        try:
            data = dict(zip(headers, row))
            
            reg_number = safe_str(data.get('VEHICLE NUMBER') or data.get('Vehicle Number') or data.get('Registration Number'))
            if not reg_number:
                continue
            
            existing = Vehicle.query.filter_by(tenant_id=tenant_id, registration_number=reg_number).first()
            if existing:
                print(f"  Skipping (exists): {reg_number}")
                continue
            
            vehicle = Vehicle(
                tenant_id=tenant_id,
                registration_number=reg_number,
                vehicle_type=safe_str(data.get('TYPE') or data.get('Type'), 50) or 'Truck',
                make=safe_str(data.get('MAKE') or data.get('Make'), 100),
                model=safe_str(data.get('MODEL') or data.get('Model'), 100),
                year=parse_int(data.get('YEAR') or data.get('Year')),
                owner_name=safe_str(data.get('OWNER NAME') or data.get('Owner Name'), 200),
                owner_contact=safe_str(data.get('OWNER CONTACT') or data.get('Owner Contact'), 50),
                load_capacity=safe_str(data.get('LOAD CAPACITY') or data.get('Truck Size'), 50),
                insurance_expiry=parse_date(data.get('INSURANCE EXPIRY') or data.get('Insurance Expiry')),
                fitness_expiry=parse_date(data.get('FITNESS EXPIRY') or data.get('Fitness Expiry')),
                status='active',
            )
            
            db.session.add(vehicle)
            imported += 1
            print(f"  Imported: {reg_number}")
            
        except Exception as e:
            print(f"  ERROR Row {row_idx}: {e}")
    
    db.session.commit()
    print(f"\nVehicles imported: {imported}")
    return imported

def import_loans(filepath, tenant_id):
    print(f"\n{'='*60}")
    print(f"Importing Loans from: {filepath}")
    print(f"{'='*60}")
    
    wb = load_workbook(filepath)
    loan_sheet = None
    
    for sheet_name in wb.sheetnames:
        if 'loan' in sheet_name.lower():
            loan_sheet = wb[sheet_name]
            break
    
    if not loan_sheet and len(wb.sheetnames) > 1:
        loan_sheet = wb[wb.sheetnames[1]]
    
    if not loan_sheet:
        print("No loan sheet found")
        return 0
    
    headers = [cell.value for cell in loan_sheet[1]]
    print(f"Sheet: {loan_sheet.title}")
    print(f"Headers: {headers}")
    
    imported = 0
    
    for row_idx, row in enumerate(loan_sheet.iter_rows(min_row=2, max_row=loan_sheet.max_row, values_only=True), start=2):
        try:
            data = dict(zip(headers, row))
            
            vehicle_number = safe_str(data.get('VEHICLE NUMBER') or data.get('Vehicle Number'))
            lender_name = safe_str(data.get('LENDER NAME') or data.get('Lender') or data.get('Bank Name'))
            
            if not vehicle_number and not lender_name:
                continue
            
            vehicle = None
            if vehicle_number:
                vehicle = Vehicle.query.filter_by(tenant_id=tenant_id, registration_number=vehicle_number).first()
            
            principal = safe_float(data.get('PRINCIPAL AMOUNT') or data.get('Principal') or data.get('Loan Amount'))
            
            existing_count = Loan.query.filter_by(tenant_id=tenant_id).count()
            loan_name = f"LOAN-{existing_count + imported + 1:05d}"
            
            loan = Loan(
                tenant_id=tenant_id,
                name=loan_name,
                vehicle_id=vehicle.id if vehicle else None,
                loan_type=safe_str(data.get('LOAN TYPE') or data.get('Loan Type'), 50) or 'Vehicle Loan',
                lender_name=lender_name or 'Unknown Lender',
                principal_amount=principal,
                interest_rate=safe_float(data.get('INTEREST RATE') or data.get('Interest Rate')),
                emi_amount=safe_float(data.get('EMI AMOUNT') or data.get('EMI')),
                total_loan_amount=principal,
                remaining_balance=principal,
                status='active',
            )
            
            db.session.add(loan)
            imported += 1
            print(f"  Imported: {loan_name} - {lender_name or 'N/A'} (₹{principal:,.2f})")
            
        except Exception as e:
            print(f"  ERROR Row {row_idx}: {e}")
    
    db.session.commit()
    print(f"\nLoans imported: {imported}")
    return imported

def import_rate_list(filepath, tenant_id):
    print(f"\n{'='*60}")
    print(f"Importing Rate List from: {filepath}")
    print(f"{'='*60}")
    
    wb = load_workbook(filepath)
    
    imported = 0
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        if ws.max_row < 2:
            continue
        
        headers = [cell.value for cell in ws[1]]
        print(f"\nSheet: {sheet_name}")
        print(f"Headers: {headers}")
        
        truck_size = sheet_name.strip() if any(x in sheet_name.lower() for x in ['wheel', 'tyre', 'truck', '10', '12', '14', '16', '18', '20']) else None
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
            try:
                data = dict(zip(headers, row))
                
                location_name = safe_str(data.get('LOCATION') or data.get('Location') or data.get('DESTINATION') or data.get('To'))
                if not location_name:
                    continue
                
                rate = safe_float(data.get('RATE') or data.get('Rate') or data.get('PRICE') or data.get('Price') or data.get('FREIGHT'))
                distance = safe_float(data.get('DISTANCE') or data.get('Distance') or data.get('KM'))
                
                city = safe_str(data.get('CITY') or data.get('City'), 100)
                state = safe_str(data.get('STATE') or data.get('State'), 100)
                
                full_location = location_name
                if truck_size and truck_size not in location_name:
                    full_location = f"{location_name} ({truck_size})"
                
                existing = Location.query.filter_by(tenant_id=tenant_id, location=full_location).first()
                if existing:
                    print(f"  Skipping (exists): {full_location}")
                    continue
                
                location = Location(
                    tenant_id=tenant_id,
                    location=full_location,
                    city=city or location_name,
                    state=state,
                    distance_km=distance if distance > 0 else None,
                    rate=rate if rate > 0 else None,
                    remarks=truck_size,
                )
                
                db.session.add(location)
                imported += 1
                if rate > 0:
                    print(f"  Imported: {full_location} - ₹{rate:,.2f}")
                else:
                    print(f"  Imported: {full_location}")
                
            except Exception as e:
                print(f"  ERROR Sheet {sheet_name}, Row {row_idx}: {e}")
    
    db.session.commit()
    print(f"\nRate list imported: {imported}")
    return imported

# Main execution
print("="*60)
print("TRANSPORT MANAGEMENT SYSTEM - EXCEL DATA IMPORT")
print("="*60)

with app.app_context():
    tenant = get_or_create_tenant()
    tenant_id = tenant.id
    print(f"\nUsing Tenant: {tenant.name} (ID: {tenant_id})")
    
    stats = {}
    
    # Import Vendors
    if os.path.exists('Vendor Details.xlsx'):
        stats['vendors'] = import_vendors('Vendor Details.xlsx', tenant_id)
    else:
        print("Vendor Details.xlsx not found")
        stats['vendors'] = 0
    
    # Import Vehicles
    if os.path.exists('Vehicle and Loan Details.xlsx'):
        stats['vehicles'] = import_vehicles('Vehicle and Loan Details.xlsx', tenant_id)
        stats['loans'] = import_loans('Vehicle and Loan Details.xlsx', tenant_id)
    else:
        print("Vehicle and Loan Details.xlsx not found")
        stats['vehicles'] = 0
        stats['loans'] = 0
    
    # Import Rate List
    if os.path.exists('ALL PARTY RATE FILE.xlsx'):
        stats['rate_list'] = import_rate_list('ALL PARTY RATE FILE.xlsx', tenant_id)
    else:
        print("ALL PARTY RATE FILE.xlsx not found")
        stats['rate_list'] = 0
    
    # Summary
    print("\n" + "="*60)
    print("IMPORT COMPLETE - FINAL SUMMARY")
    print("="*60)
    for module, count in stats.items():
        print(f"  {module.capitalize():12}: {count} imported")

print(f"\nCompleted at: {datetime.now()}")

# Restore stdout/stderr
sys.stdout = original_stdout
sys.stderr = original_stderr
log_file.close()

print("Import completed! Check import_log.txt for details.")
