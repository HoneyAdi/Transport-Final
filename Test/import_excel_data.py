"""
Import data from Excel files into the Transport Management System
"""
import os
import sys
from datetime import datetime
from openpyxl import load_workbook

# Add the project directory to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from models import db, Vendor, Vehicle, Loan, Location, Tenant, app
from webapp import generate_vendor_code

def parse_date(date_value):
    """Parse date from various formats"""
    if not date_value:
        return None
    if isinstance(date_value, datetime):
        return date_value.date()
    if isinstance(date_value, str):
        try:
            return datetime.strptime(date_value, '%Y-%m-%d').date()
        except:
            try:
                return datetime.strptime(date_value, '%d/%m/%Y').date()
            except:
                return None
    return None

def safe_str(value, max_len=None):
    """Safely convert to string"""
    if value is None:
        return None
    result = str(value).strip()
    if max_len and len(result) > max_len:
        result = result[:max_len]
    return result if result else None

def safe_float(value):
    """Safely convert to float"""
    if value is None:
        return 0.0
    try:
        if isinstance(value, str):
            value = value.replace(',', '').replace('₹', '').strip()
        return float(value)
    except:
        return 0.0

def get_or_create_tenant():
    """Get first tenant or create one"""
    tenant = Tenant.query.first()
    if not tenant:
        tenant = Tenant(name="Default Tenant", subdomain="default")
        db.session.add(tenant)
        db.session.commit()
        print(f"Created tenant: {tenant.name} (ID: {tenant.id})")
    return tenant

def import_vendors_from_excel(filepath, tenant_id):
    """Import vendors from Excel file"""
    print(f"\n{'='*60}")
    print(f"Importing Vendors from: {filepath}")
    print(f"{'='*60}")
    
    wb = load_workbook(filepath)
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    print(f"Headers: {headers}")
    print(f"Total rows: {ws.max_row - 1}")
    
    imported = 0
    errors = []
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
        try:
            data = dict(zip(headers, row))
            
            # Skip if no vendor name
            vendor_name = safe_str(data.get('VENDOR NAME') or data.get('Company Name') or data.get('Vendor Name'))
            if not vendor_name:
                continue
            
            # Check if vendor already exists
            existing = Vendor.query.filter_by(tenant_id=tenant_id, vendor_name=vendor_name).first()
            if existing:
                print(f"  Skipping (exists): {vendor_name}")
                continue
            
            # Generate vendor code
            vendor_code = generate_vendor_code(tenant_id)
            
            vendor = Vendor(
                tenant_id=tenant_id,
                vendor_code=vendor_code,
                vendor_name=vendor_name,
                vendor_type=safe_str(data.get('TYPE') or data.get('Vendor Type') or data.get('vendor_type'), 50) or 'supplier',
                status='active',
                contact_person=safe_str(data.get('CONTACT PERSON') or data.get('Contact Person') or data.get('contact_person'), 100),
                designation=safe_str(data.get('DESIGNATION') or data.get('Designation'), 100),
                phone_primary=safe_str(data.get('PHONE') or data.get('Phone') or data.get('phone_primary'), 20),
                mobile=safe_str(data.get('MOBILE') or data.get('Mobile') or data.get('mobile'), 20),
                email=safe_str(data.get('EMAIL') or data.get('Email') or data.get('email'), 150),
                reg_address_line1=safe_str(data.get('ADDRESS') or data.get('Address') or data.get('reg_address_line1'), 200),
                reg_city=safe_str(data.get('CITY') or data.get('City') or data.get('reg_city'), 100),
                reg_state=safe_str(data.get('STATE') or data.get('State') or data.get('reg_state'), 100),
                reg_pincode=safe_str(data.get('PINCODE') or data.get('Pincode') or data.get('reg_pincode'), 10),
                reg_country='India',
                business_nature=safe_str(data.get('BUSINESS NATURE') or data.get('Business Nature'), 100),
                website=safe_str(data.get('WEBSITE') or data.get('Website'), 200),
                gstin=safe_str(data.get('GSTIN') or data.get('Gstin') or data.get('GST'), 20),
                pan=safe_str(data.get('PAN') or data.get('Pan'), 20),
                tan=safe_str(data.get('TAN') or data.get('Tan'), 20),
                bank_name=safe_str(data.get('BANK NAME') or data.get('Bank Name'), 100),
                bank_account_number=safe_str(data.get('ACCOUNT NUMBER') or data.get('Account Number'), 50),
                bank_ifsc=safe_str(data.get('IFSC') or data.get('IFSC Code'), 20),
                bank_branch=safe_str(data.get('BRANCH') or data.get('Branch'), 100),
            )
            
            db.session.add(vendor)
            imported += 1
            print(f"  Imported: {vendor_name} (Code: {vendor_code})")
            
        except Exception as e:
            error_msg = f"Row {row_idx}: {str(e)}"
            errors.append(error_msg)
            print(f"  ERROR: {error_msg}")
    
    db.session.commit()
    print(f"\nVendor Import Summary:")
    print(f"  Successfully imported: {imported}")
    print(f"  Errors: {len(errors)}")
    return imported, errors

def import_vehicles_from_excel(filepath, tenant_id):
    """Import vehicles from Excel file"""
    print(f"\n{'='*60}")
    print(f"Importing Vehicles from: {filepath}")
    print(f"{'='*60}")
    
    wb = load_workbook(filepath)
    
    # Look for Vehicle sheet
    vehicle_sheet = None
    for sheet_name in wb.sheetnames:
        if 'vehicle' in sheet_name.lower():
            vehicle_sheet = wb[sheet_name]
            break
    
    if not vehicle_sheet and len(wb.sheetnames) > 0:
        vehicle_sheet = wb[wb.sheetnames[0]]  # Use first sheet
    
    if not vehicle_sheet:
        print("No vehicle sheet found!")
        return 0, []
    
    headers = [cell.value for cell in vehicle_sheet[1]]
    print(f"Sheet: {vehicle_sheet.title}")
    print(f"Headers: {headers}")
    print(f"Total rows: {vehicle_sheet.max_row - 1}")
    
    imported = 0
    errors = []
    
    for row_idx, row in enumerate(vehicle_sheet.iter_rows(min_row=2, max_row=vehicle_sheet.max_row, values_only=True), start=2):
        try:
            data = dict(zip(headers, row))
            
            # Get vehicle number
            reg_number = safe_str(data.get('VEHICLE NUMBER') or data.get('Vehicle Number') or data.get('Registration Number') or data.get('reg_number'))
            if not reg_number:
                continue
            
            # Check if exists
            existing = Vehicle.query.filter_by(tenant_id=tenant_id, registration_number=reg_number).first()
            if existing:
                print(f"  Skipping (exists): {reg_number}")
                continue
            
            vehicle = Vehicle(
                tenant_id=tenant_id,
                registration_number=reg_number,
                vehicle_type=safe_str(data.get('TYPE') or data.get('Type') or data.get('vehicle_type'), 50) or 'Truck',
                make=safe_str(data.get('MAKE') or data.get('Make'), 100),
                model=safe_str(data.get('MODEL') or data.get('Model'), 100),
                year=parse_int(data.get('YEAR') or data.get('Year')),
                color=safe_str(data.get('COLOR') or data.get('Color'), 50),
                owner_name=safe_str(data.get('OWNER NAME') or data.get('Owner Name') or data.get('owner_name'), 200),
                owner_contact=safe_str(data.get('OWNER CONTACT') or data.get('Owner Contact') or data.get('owner_contact'), 50),
                chassis_number=safe_str(data.get('CHASSIS NUMBER') or data.get('Chassis Number'), 100),
                engine_number=safe_str(data.get('ENGINE NUMBER') or data.get('Engine Number'), 100),
                fuel_type=safe_str(data.get('FUEL TYPE') or data.get('Fuel Type'), 50),
                seating_capacity=parse_int(data.get('SEATING CAPACITY') or data.get('Seating Capacity')),
                load_capacity=safe_str(data.get('LOAD CAPACITY') or data.get('Load Capacity') or data.get('TRUCK SIZE'), 50),
                insurance_expiry=parse_date(data.get('INSURANCE EXPIRY') or data.get('Insurance Expiry')),
                fitness_expiry=parse_date(data.get('FITNESS EXPIRY') or data.get('Fitness Expiry')),
                permit_expiry=parse_date(data.get('PERMIT EXPIRY') or data.get('Permit Expiry')),
                pollution_expiry=parse_date(data.get('POLLUTION EXPIRY') or data.get('Pollution Expiry')),
                tax_expiry=parse_date(data.get('TAX EXPIRY') or data.get('Tax Expiry')),
                status='active',
            )
            
            db.session.add(vehicle)
            imported += 1
            print(f"  Imported: {reg_number}")
            
        except Exception as e:
            error_msg = f"Row {row_idx}: {str(e)}"
            errors.append(error_msg)
            print(f"  ERROR: {error_msg}")
    
    db.session.commit()
    print(f"\nVehicle Import Summary:")
    print(f"  Successfully imported: {imported}")
    print(f"  Errors: {len(errors)}")
    return imported, errors

def import_loans_from_excel(filepath, tenant_id):
    """Import loans from Excel file"""
    print(f"\n{'='*60}")
    print(f"Importing Loans from: {filepath}")
    print(f"{'='*60}")
    
    wb = load_workbook(filepath)
    
    # Look for Loan sheet
    loan_sheet = None
    for sheet_name in wb.sheetnames:
        if 'loan' in sheet_name.lower():
            loan_sheet = wb[sheet_name]
            break
    
    if not loan_sheet and len(wb.sheetnames) > 1:
        loan_sheet = wb[wb.sheetnames[1]]  # Use second sheet if no loan sheet found
    
    if not loan_sheet:
        print("No loan sheet found!")
        return 0, []
    
    headers = [cell.value for cell in loan_sheet[1]]
    print(f"Sheet: {loan_sheet.title}")
    print(f"Headers: {headers}")
    print(f"Total rows: {loan_sheet.max_row - 1}")
    
    imported = 0
    errors = []
    
    for row_idx, row in enumerate(loan_sheet.iter_rows(min_row=2, max_row=loan_sheet.max_row, values_only=True), start=2):
        try:
            data = dict(zip(headers, row))
            
            # Get loan details
            vehicle_number = safe_str(data.get('VEHICLE NUMBER') or data.get('Vehicle Number'))
            lender_name = safe_str(data.get('LENDER NAME') or data.get('Lender') or data.get('Lender Name') or data.get('BANK NAME') or data.get('Bank Name'))
            
            if not vehicle_number and not lender_name:
                continue
            
            # Find vehicle
            vehicle = None
            if vehicle_number:
                vehicle = Vehicle.query.filter_by(tenant_id=tenant_id, registration_number=vehicle_number).first()
            
            principal = safe_float(data.get('PRINCIPAL AMOUNT') or data.get('Principal') or data.get('Loan Amount') or data.get('PRINCIPAL'))
            
            # Generate loan name
            existing_count = Loan.query.filter_by(tenant_id=tenant_id).count()
            loan_name = f"LOAN-{existing_count + imported + 1:05d}"
            
            loan = Loan(
                tenant_id=tenant_id,
                name=loan_name,
                vehicle_id=vehicle.id if vehicle else None,
                loan_type=safe_str(data.get('LOAN TYPE') or data.get('Loan Type'), 50) or 'Vehicle Loan',
                lender_name=lender_name or 'Unknown Lender',
                principal_amount=principal,
                interest_rate=safe_float(data.get('INTEREST RATE') or data.get('Interest Rate')),
                emi_amount=safe_float(data.get('EMI AMOUNT') or data.get('EMI')),
                total_loan_amount=principal,
                remaining_balance=principal,
                status='active',
            )
            
            db.session.add(loan)
            imported += 1
            print(f"  Imported: {loan_name} - {lender_name or 'N/A'} (₹{principal:,.2f})")
            
        except Exception as e:
            error_msg = f"Row {row_idx}: {str(e)}"
            errors.append(error_msg)
            print(f"  ERROR: {error_msg}")
    
    db.session.commit()
    print(f"\nLoan Import Summary:")
    print(f"  Successfully imported: {imported}")
    print(f"  Errors: {len(errors)}")
    return imported, errors

def import_rate_list_from_excel(filepath, tenant_id):
    """Import rate list from Excel file with truck size-based rates"""
    print(f"\n{'='*60}")
    print(f"Importing Rate List from: {filepath}")
    print(f"{'='*60}")
    
    wb = load_workbook(filepath)
    
    imported = 0
    errors = []
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        if ws.max_row < 2:
            continue
        
        headers = [cell.value for cell in ws[1]]
        print(f"\nSheet: {sheet_name}")
        print(f"Headers: {headers}")
        print(f"Total rows: {ws.max_row - 1}")
        
        # Detect truck size from sheet name or headers
        truck_size = None
        if any(size in sheet_name.lower() for size in ['10', '12', '14', '16', '18', '20', '22', '24', 'wheeler', 'tyre']):
            # Extract truck size from sheet name
            truck_size = sheet_name.strip()
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
            try:
                data = dict(zip(headers, row))
                
                # Get location details
                location_name = safe_str(data.get('LOCATION') or data.get('Location') or data.get('DESTINATION') or data.get('Destination') or data.get('TO') or data.get('To'))
                if not location_name:
                    continue
                
                # Get rate - could be different column names
                rate = safe_float(data.get('RATE') or data.get('Rate') or data.get('PRICE') or data.get('Price') or data.get('FREIGHT') or data.get('Freight'))
                
                # Get distance if available
                distance = safe_float(data.get('DISTANCE') or data.get('Distance') or data.get('KM') or data.get('KMS'))
                
                # Get city/state if available
                city = safe_str(data.get('CITY') or data.get('City'), 100)
                state = safe_str(data.get('STATE') or data.get('State'), 100)
                
                # Create location name with truck size info
                full_location = location_name
                if truck_size and truck_size not in location_name:
                    full_location = f"{location_name} ({truck_size})"
                
                # Check if exists
                existing = Location.query.filter_by(tenant_id=tenant_id, location=full_location).first()
                if existing:
                    print(f"  Skipping (exists): {full_location}")
                    continue
                
                location = Location(
                    tenant_id=tenant_id,
                    location=full_location,
                    city=city or location_name,
                    state=state,
                    distance_km=distance if distance > 0 else None,
                    rate=rate if rate > 0 else None,
                    remarks=safe_str(data.get('REMARKS') or data.get('Remarks') or data.get('NOTE') or truck_size, 500),
                )
                
                db.session.add(location)
                imported += 1
                print(f"  Imported: {full_location} - ₹{rate:,.2f}" if rate > 0 else f"  Imported: {full_location}")
                
            except Exception as e:
                error_msg = f"Sheet {sheet_name}, Row {row_idx}: {str(e)}"
                errors.append(error_msg)
                print(f"  ERROR: {error_msg}")
    
    db.session.commit()
    print(f"\nRate List Import Summary:")
    print(f"  Successfully imported: {imported}")
    print(f"  Errors: {len(errors)}")
    return imported, errors

def parse_int(value):
    """Safely parse integer"""
    if value is None:
        return None
    try:
        return int(float(value))
    except:
        return None

def main():
    print("="*60)
    print("TRANSPORT MANAGEMENT SYSTEM - EXCEL DATA IMPORT")
    print("="*60)
    
    with app.app_context():
        # Get or create tenant
        tenant = get_or_create_tenant()
        tenant_id = tenant.id
        print(f"\nUsing Tenant: {tenant.name} (ID: {tenant_id})")
        
        total_stats = {
            'vendors': (0, []),
            'vehicles': (0, []),
            'loans': (0, []),
            'rate_list': (0, [])
        }
        
        # Import Vendors
        vendor_file = 'Vendor Details.xlsx'
        if os.path.exists(vendor_file):
            total_stats['vendors'] = import_vendors_from_excel(vendor_file, tenant_id)
        else:
            print(f"\nVendor file not found: {vendor_file}")
        
        # Import Vehicles
        vehicle_file = 'Vehicle and Loan Details.xlsx'
        if os.path.exists(vehicle_file):
            total_stats['vehicles'] = import_vehicles_from_excel(vehicle_file, tenant_id)
        else:
            print(f"\nVehicle file not found: {vehicle_file}")
        
        # Import Loans
        if os.path.exists(vehicle_file):
            total_stats['loans'] = import_loans_from_excel(vehicle_file, tenant_id)
        
        # Import Rate List
        rate_file = 'ALL PARTY RATE FILE.xlsx'
        if os.path.exists(rate_file):
            total_stats['rate_list'] = import_rate_list_from_excel(rate_file, tenant_id)
        else:
            print(f"\nRate list file not found: {rate_file}")
        
        # Final Summary
        print("\n" + "="*60)
        print("IMPORT COMPLETE - FINAL SUMMARY")
        print("="*60)
        for module, (count, errors) in total_stats.items():
            print(f"  {module.capitalize():12}: {count} imported, {len(errors)} errors")
        
        total_errors = sum(len(errors) for _, errors in total_stats.values())
        if total_errors > 0:
            print(f"\nTotal errors: {total_errors}")
            print("Check the output above for specific error details.")

if __name__ == '__main__':
    main()
